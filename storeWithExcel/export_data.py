import json
from openpyxl import Workbook

# Carrega os dados do arquivo JSON
with open('data.json', 'r') as json_file:
    data = json.load(json_file)

# Cria uma nova planilha Excel
wb = Workbook()

# Função para preencher uma planilha com os dados fornecidos
def populate_sheet(sheet, header, rows):
    sheet.append(header)
    for row in rows:
        sheet.append(row)

# Preenche a planilha de Clientes
populate_sheet(wb.active, ['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score'],
               [[client['name'], client['date'], client['cpf'], client['origin'], client['score']]
                for client in data['clients']])

# Preenche a planilha de Vendedores
populate_sheet(wb.create_sheet(title='Vendedores'), ['Nome', 'Matrícula'],
               [[seller['name'], seller['matricula']] for seller in data['sellers']])

# Preenche a planilha de Produtos
populate_sheet(wb.create_sheet(title='Produtos'), ['Nome', 'Valor', 'Categoria'],
               [[product['name'], product['value'], product['category']] for product in data['products']])

# Salva a planilha Excel com os dados
wb.save('data.xlsx')
print('Planilha Excel gerada com sucesso!')
